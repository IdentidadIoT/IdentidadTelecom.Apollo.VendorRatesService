"""
Infraestructura base para lectura de archivos Excel basada en configuración.

Este módulo proporciona componentes genéricos y reutilizables para leer archivos Excel
de diferentes vendors sin duplicación de código.
"""
from dataclasses import dataclass, field
from typing import Dict, Callable, Any, List, Optional
import openpyxl
from core.logging import logger


@dataclass
class SheetConfig:
    """
    Configuración para leer una hoja de Excel específica.

    Attributes:
        name: Nombre de la hoja en el archivo Excel
        start_row: Fila desde la cual empezar a leer datos (1-indexed)
        column_mapping: Diccionario que mapea nombres de campos a índices de columna (0-indexed)
        transformations: Funciones de transformación opcionales para campos específicos
        fallback_sheet: Nombre alternativo de hoja o "FIRST" para usar la primera hoja
    """
    name: str
    start_row: int
    column_mapping: Dict[str, int]
    transformations: Optional[Dict[str, Callable]] = None
    fallback_sheet: Optional[str] = None


@dataclass
class VendorExcelConfig:
    """
    Configuración completa de lectura Excel para un vendor.

    Attributes:
        vendor_name: Nombre del vendor (para logging)
        sheets: Diccionario que mapea tipos de hoja a su configuración
    """
    vendor_name: str
    sheets: Dict[str, SheetConfig]


class ExcelReaderBase:
    """
    Lector genérico de archivos Excel basado en configuración.

    Esta clase elimina la duplicación de código al proporcionar un método genérico
    que puede leer cualquier estructura de Excel basándose en una configuración declarativa.
    """

    @staticmethod
    def read_sheet(
        file_path: str,
        config: SheetConfig,
        vendor_name: str
    ) -> List[Dict[str, Any]]:
        """
        Lee una hoja de Excel basándose en la configuración proporcionada.

        Args:
            file_path: Ruta al archivo Excel
            config: Configuración de la hoja a leer
            vendor_name: Nombre del vendor (para logging)

        Returns:
            Lista de diccionarios con los datos leídos

        Raises:
            Exception: Si ocurre un error al leer el archivo
        """
        try:
            # Abrir workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Buscar hoja (con fallback si está configurado)
            sheet = ExcelReaderBase._find_sheet(workbook, config)
            if not sheet:
                logger.error(f"[{vendor_name}] Hoja '{config.name}' no encontrada en {file_path}")
                workbook.close()
                return []

            # Leer filas - IGUAL que código C# (usa LastRowUsed, lee hasta el final sin parar)
            # C#: int lastRow = worksheet.LastRowUsed().RowNumber();
            #     for (int i = 9; i <= lastRow; i++) { ... }
            last_row = sheet.max_row
            logger.debug(f"[{vendor_name}] {config.name}: Leyendo filas {config.start_row} a {last_row}")

            data = []
            for row in sheet.iter_rows(min_row=config.start_row, max_row=last_row, values_only=True):
                # Mapear columnas a campos
                item = {}
                for field_name, col_idx in config.column_mapping.items():
                    # Obtener valor de la celda (si existe)
                    value = row[col_idx] if col_idx >= 0 and len(row) > col_idx else None

                    # Aplicar transformación si existe
                    if config.transformations and field_name in config.transformations:
                        try:
                            value = config.transformations[field_name](value, row)
                        except Exception as transform_error:
                            logger.warning(
                                f"[{vendor_name}] Error transformando campo '{field_name}': {transform_error}"
                            )
                            value = ""
                    else:
                        # Transformación default: convertir a string y limpiar
                        value = str(value).strip() if value is not None else ""

                    item[field_name] = value

                data.append(item)

            workbook.close()
            logger.info(f"[{vendor_name}] {config.name}: {len(data)} registros leídos")
            return data

        except Exception as e:
            logger.error(f"[{vendor_name}] Error leyendo hoja '{config.name}': {e}", exc_info=True)
            raise

    @staticmethod
    def _find_sheet(workbook, config: SheetConfig):
        """
        Busca una hoja en el workbook con soporte para fallback.

        Args:
            workbook: Workbook de openpyxl
            config: Configuración de la hoja

        Returns:
            Worksheet encontrada o None
        """
        # Intentar con nombre principal
        if config.name in workbook.sheetnames:
            return workbook[config.name]

        # Intentar con fallback
        if config.fallback_sheet:
            if config.fallback_sheet == "FIRST":
                # Usar primera hoja del workbook
                return workbook.worksheets[0]
            elif config.fallback_sheet in workbook.sheetnames:
                return workbook[config.fallback_sheet]

        return None
